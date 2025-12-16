import openpyxl
import csv
import sys

def parse_score_range(score_str):
    if isinstance(score_str, (int, float)):
        return float(score_str), float(score_str)
    elif isinstance(score_str, str) and '-' in score_str:
        min_str, max_str = score_str.split('-')
        return float(min_str.strip()), float(max_str.strip())
    else:
        raise ValueError(f"Invalid score range: {score_str}")

def main(input_file, output_file):
    wb = openpyxl.load_workbook(input_file, data_only=True)
    sheet = wb.active

    khoi_list = ['A', 'A1', 'B', 'C', 'D']
    years_count = 13  # 2025 to 2013

    with open(output_file, 'w', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(['max_score', 'min_score', 'year', 'khoi', 'count', 'cumulative'])

        row = 1
        while row <= sheet.max_row:
            cell_value = sheet.cell(row, 1).value
            if cell_value in khoi_list:
                khoi = cell_value
                years = []
                for col in range(2, 2 + years_count):
                    year = sheet.cell(row, col).value
                    if year:
                        years.append(year)
                if len(years) != years_count:
                    raise ValueError(f"Unexpected number of years for khoi {khoi}")

                data_rows = []
                r = row + 1
                while r <= sheet.max_row:
                    score_str = sheet.cell(r, 1).value
                    if score_str is None or (isinstance(score_str, str) and (score_str.startswith('Tổng') or score_str.startswith('Điểm'))):
                        break

                    try:
                        max_sc, min_sc = parse_score_range(score_str)  # Note: max first as per user (highest in range)
                    except ValueError:
                        break  # Skip if not parsable

                    counts = []
                    for col in range(2, 2 + years_count):
                        cnt = sheet.cell(r, col).value
                        if cnt is None:
                            cnt = 0
                        counts.append(int(cnt))  # Assume counts are integers

                    data_rows.append((max_sc, min_sc, counts))
                    r += 1

                # Compute cumulatives for each year
                num_ranges = len(data_rows)
                if num_ranges > 0:
                    cumuls = [[0] * num_ranges for _ in range(years_count)]
                    for y_idx in range(years_count):
                        current_sum = 0
                        for rng_idx in range(num_ranges):
                            current_sum += data_rows[rng_idx][2][y_idx]
                            cumuls[y_idx][rng_idx] = current_sum

                    # Write to CSV
                    for rng_idx in range(num_ranges):
                        max_sc, min_sc, counts = data_rows[rng_idx]
                        for y_idx, year in enumerate(years):
                            count = counts[y_idx]
                            cumul = cumuls[y_idx][rng_idx]
                            writer.writerow([max_sc, min_sc, year, khoi, count, cumul])

                # Skip to next section
                row = r
                while row <= sheet.max_row:
                    next_value = sheet.cell(row, 1).value
                    if next_value in khoi_list or next_value is None:
                        break
                    row += 1
            else:
                row += 1

if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python script.py input.xlsx output.csv")
        sys.exit(1)
    main(sys.argv[1], sys.argv[2])